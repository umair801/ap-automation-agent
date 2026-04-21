# exports/excel_exporter.py

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from agents.extraction_agent import ScheduleExtractionResult, ScheduleItem
from core.logger import get_logger

logger = get_logger(__name__)

# ─── Default column mapping ───────────────────────────────────────────────────
# Keys are the Excel header labels.
# Values are the ScheduleItem attribute names.
# Reorder or rename keys to match any client's takeoff format.

DEFAULT_COLUMN_MAP: dict[str, str] = {
    "Type Mark":          "type_mark",
    "Item Type":          "item_type",
    "Size":               "size",
    "Width (mm)":         "width_mm",
    "Height (mm)":        "height_mm",
    "Quantity":           "quantity",
    "Location":           "location",
    "Room Ref":           "room_reference",
    "Elevation Ref":      "elevation_reference",
    "Finish":             "finish",
    "Frame Type":         "frame_type",
    "Glazing":            "glazing",
    "Notes":              "notes",
}

# ─── Style constants ──────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ALT_ROW_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
TITLE_FONT    = Font(name="Calibri", bold=True, size=13)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass
class ExportConfig:
    """Controls how the Excel file is generated."""
    output_path: str                         # Full path including filename
    column_map: dict[str, str] = field(default_factory=lambda: dict(DEFAULT_COLUMN_MAP))
    sheet_title: str = "Window & Door Schedule"
    include_page_column: bool = True         # Adds a 'Source Page' column
    confidence_threshold: float = 0.5        # Skip pages below this score
    only_passed: bool = True                 # If True, skip failed pages


@dataclass
class ExportResult:
    output_path: str
    total_rows: int
    pages_included: int
    pages_skipped: int
    success: bool
    error: Optional[str] = None


def _get_cell_value(item: ScheduleItem, attr: str):
    """Safely retrieve an attribute value from a ScheduleItem."""
    return getattr(item, attr, None)


def _apply_header_row(ws, headers: list[str], row: int = 1):
    """Write and style the header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _apply_body_row(ws, values: list, row: int, alternate: bool):
    """Write and style a single data row."""
    fill = ALT_ROW_FILL if alternate else None
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _set_column_widths(ws, headers: list[str]):
    """Set reasonable column widths based on header name length."""
    width_map = {
        "Type Mark": 12, "Item Type": 12, "Size": 14,
        "Width (mm)": 12, "Height (mm)": 13, "Quantity": 10,
        "Location": 22, "Room Ref": 12, "Elevation Ref": 16,
        "Finish": 18, "Frame Type": 14, "Glazing": 20,
        "Notes": 30, "Source Page": 13,
    }
    for col_idx, header in enumerate(headers, start=1):
        width = width_map.get(header, max(len(header) + 4, 12))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def export_to_excel(
    results: list[ScheduleExtractionResult],
    config: ExportConfig,
) -> ExportResult:
    """
    Write extracted schedule items to a formatted Excel file.

    Args:
        results: List of ScheduleExtractionResult from the extraction agent.
        config:  ExportConfig controlling output path, columns, and filters.

    Returns:
        ExportResult with row count, page stats, and success flag.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config.sheet_title[:31]   # Excel tab name limit is 31 chars
    ws.freeze_panes = "A3"               # Freeze title + header rows

    # ── Title row ────────────────────────────────────────────────────────────
    headers = list(config.column_map.keys())
    if config.include_page_column:
        headers.append("Source Page")

    title_cell = ws.cell(row=1, column=1, value=config.sheet_title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = LEFT_ALIGN
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=len(headers)
    )

    # ── Sub-title with export timestamp ──────────────────────────────────────
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws.cell(row=1, column=len(headers) + 1, value=f"Exported: {ts}").font = Font(
        name="Calibri", italic=True, size=9, color="666666"
    )

    # ── Header row ────────────────────────────────────────────────────────────
    _apply_header_row(ws, headers, row=2)
    ws.row_dimensions[2].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_row = 3
    total_rows = 0
    pages_included = 0
    pages_skipped = 0

    for result in results:
        # Apply filters
        if config.only_passed and not result.passed_threshold:
            logger.info(
                "Skipping page below threshold",
                page=result.page_number,
                confidence=result.confidence,
            )
            pages_skipped += 1
            continue

        if result.confidence < config.confidence_threshold:
            pages_skipped += 1
            continue

        pages_included += 1

        for idx, item in enumerate(result.items):
            values = [
                _get_cell_value(item, attr)
                for attr in config.column_map.values()
            ]
            if config.include_page_column:
                values.append(result.page_number + 1)   # 1-indexed for users

            alternate = (total_rows % 2 == 1)
            _apply_body_row(ws, values, current_row, alternate)
            current_row += 1
            total_rows += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    _set_column_widths(ws, headers)

    # ── Summary row ───────────────────────────────────────────────────────────
    if total_rows > 0:
        summary_row = current_row + 1
        summary_cell = ws.cell(
            row=summary_row,
            column=1,
            value=f"Total items: {total_rows}  |  Pages processed: {pages_included}  |  Pages skipped: {pages_skipped}",
        )
        summary_cell.font = Font(name="Calibri", italic=True, size=9, color="444444")

    # ── Save ──────────────────────────────────────────────────────────────────
    output_path = Path(config.output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(str(output_path))
        logger.info(
            "Excel export complete",
            path=str(output_path),
            rows=total_rows,
            pages_included=pages_included,
        )
        return ExportResult(
            output_path=str(output_path),
            total_rows=total_rows,
            pages_included=pages_included,
            pages_skipped=pages_skipped,
            success=True,
        )
    except Exception as e:
        logger.error("Failed to save Excel file", error=str(e))
        return ExportResult(
            output_path=str(output_path),
            total_rows=0,
            pages_included=0,
            pages_skipped=0,
            success=False,
            error=str(e),
        )